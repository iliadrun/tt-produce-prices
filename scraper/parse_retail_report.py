"""Parse NAMDEVCO's monthly retail price workbook into clean JSON.

The workbook ("Average Retail Prices for Fresh Produce at Retail Markets")
holds ONE survey month at a time — ~525-550 rows of commodity (Variety)
(Grade), priced per lb / Each / Bundle / Head / Pack at four outlet types
plus an "All Retail Markets" average column.

Layout facts this parser depends on (verified against the March and May
2026 files, which differ in content but not structure):
  - Single sheet "Avg Retail Prices"; merged title cell A1 ends with the
    survey month ("... for Mar 2026"). The month is parsed from THERE —
    filenames and link text on the source websites have both lied before.
  - Header in row 2 (ignored; columns are positional A-G).
  - Category divider rows have an empty Unit column. The SET of categories
    is not stable between months (March had FRUITS, May dropped it), so the
    category is recorded as advisory text only — nothing downstream may
    rely on which categories exist.
  - Missing prices are the literal string 'na' (any non-number → null).
  - (commodity, unit) is NOT unique: both files contain 15-20 repeated
    names with different prices. Rows are kept as a list, never a dict.
  - The "All Retail Markets" column is computed by NAMDEVCO from survey
    data we can't see and is sometimes inconsistent with the four visible
    outlet columns. It is stored as its own value, never recomputed.
  - The sheet ends with a disclaimer row merged across columns A-G; data
    rows are never merged, so a merged row below the title means stop.
"""

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

SHEET_NAME = "Avg Retail Prices"

# Column order: A=name, B=unit, then the price columns.
OUTLET_KEYS = ["farmers_markets", "municipal_markets", "vege_marts",
               "supermarkets", "all_retail"]

MONTH_NUMBERS = {}
for _i, _name in enumerate(
        ["January", "February", "March", "April", "May", "June", "July",
         "August", "September", "October", "November", "December"], start=1):
    MONTH_NUMBERS[_name.lower()] = _i
    MONTH_NUMBERS[_name[:3].lower()] = _i

TITLE_RE = re.compile(r"for\s+([A-Za-z]+)\.?,?\s+(\d{4})\s*$")


def slugify(text):
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")


def split_name(raw):
    """'TOMATO (Hybrid) (A)' -> ('TOMATO', 'Hybrid', 'A').

    The base is everything before the first '(', the variety is the first
    paren group, the grade the last (also egg sizes L/M/S). Names without
    parens are base-only.
    """
    base = raw.split("(", 1)[0].strip()
    groups = [g.strip() for g in re.findall(r"\(([^()]*)\)", raw)]
    groups = [g for g in groups if g]
    if not base:           # defensive: a name that starts with '('
        return raw.strip(), None, None
    variety = groups[0] if groups else None
    grade = groups[-1] if len(groups) > 1 else None
    return base, variety, grade


def parse_retail_report(source):
    """Parse a retail workbook (path or file-like) into a dict."""
    wb = load_workbook(source, data_only=True, read_only=False)
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        ws = wb[wb.sheetnames[0]]

    title = str(ws.cell(row=1, column=1).value or "").strip()
    m = TITLE_RE.search(title)
    if not m:
        raise ValueError(f"can't find a survey month in the title: {title!r}")
    month_num = MONTH_NUMBERS.get(m.group(1).lower())
    if not month_num:
        raise ValueError(f"unrecognised month name in title: {title!r}")
    survey_month = f"{m.group(2)}-{month_num:02d}"

    # Rows merged across the table width: the title and the disclaimer
    # footer. Hitting one below the header means the data is over.
    merged_rows = set()
    for rng in ws.merged_cells.ranges:
        if rng.min_col == 1 and rng.max_col >= len(OUTLET_KEYS):
            merged_rows.update(range(rng.min_row, rng.max_row + 1))

    items = []
    category = None
    for row in range(3, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if name is None or not str(name).strip():
            continue
        name = str(name).strip()
        if row in merged_rows or name.lower().startswith("the prices reflected"):
            break
        unit = ws.cell(row=row, column=2).value
        if unit is None or not str(unit).strip():
            category = name
            continue
        base, variety, grade = split_name(name)
        prices = {}
        for i, key in enumerate(OUTLET_KEYS):
            v = ws.cell(row=row, column=3 + i).value
            # 'na' (and anything else non-numeric) means not available.
            prices[key] = round(float(v), 4) if isinstance(v, (int, float)) else None
        items.append({
            "name_raw": name,
            "base": base,
            "base_id": slugify(base),
            "variety": variety,
            "grade": grade,
            "unit": str(unit).strip(),
            "category_raw": category,
            "prices": prices,
        })

    if not items:
        raise ValueError("no item rows found — layout may have changed")
    return {
        "survey_month": survey_month,
        "title": title,
        "currency": "TTD",
        "items": items,
    }


def main():
    if len(sys.argv) != 2:
        raise SystemExit("usage: py scraper/parse_retail_report.py <workbook.xlsx>")
    report = parse_retail_report(Path(sys.argv[1]))
    json.dump(report, sys.stdout, indent=2)


if __name__ == "__main__":
    main()
